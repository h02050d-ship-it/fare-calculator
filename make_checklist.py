# -*- coding: utf-8 -*-
# 送料計算ツール 配布チェックリスト（xlsx）を作成
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import sys

OUT = sys.argv[1] if len(sys.argv) > 1 else "送料ツール_配布チェックリスト.xlsx"
URL = "https://h02050d-ship-it.github.io/fare-calculator/soryo.html"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "配布チェックリスト"
ws.sheet_view.showGridLines = False

thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="2E5E3A")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=14, color="2E5E3A")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# タイトル
ws.merge_cells("A1:K1")
ws["A1"] = "送料計算ツール（市場用）配布チェックリスト"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A2:K2")
ws["A2"] = f"作成日 2026-06-25 ／ ツールURL：{URL} ／ 状況欄はプルダウンで更新できます"
ws["A2"].font = Font(size=9, color="666666")

headers = ["No","市場・取引先名","担当者","区分","連絡方法","メールアドレス","TEL","FAX","状況","送付日","備考"]
hrow = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=hrow, column=c, value=h)
    cell.fill = HEAD_FILL; cell.font = HEAD_FONT; cell.alignment = center; cell.border = border

rows = [
    # No, 名称, 担当, 区分, 連絡方法, メール, TEL, FAX, 状況, 送付日, 備考
    [1,"丸宇木材市売㈱ 京葉市場","桝島 康平","市場","メール","masujima@maruu.co.jp","047-442-6011","047-442-6010","下書き作成済","","送料見積を都度依頼→ツールの最適対象。下書き済（要送信）"],
    [2,"丸宇木材市売㈱ 下館市場","豊嶋 一浩","市場","メール","toyoshima@maruu.co.jp","0296-30-7001","0296-30-7015","下書き作成済","","見積依頼の常連。下書き済（要送信）"],
    [3,"㈱勝山木材市場","—","市場","メール","katsuichi-1@triton.ocn.ne.jp","","","下書き作成済","","精算書取引。下書き済（要送信）"],
    [4,"ナイス㈱ 沼津木材営業所","上柳 麻里子","卸/市場","メール","mariko.kamiyanagi@nice.co.jp","055-967-3151","055-967-3155","下書き作成済","","委託販売寄り。送付要否は要判断"],
    [5,"丸宇木材市売㈱ 本社総務部","角田 彩恵","経理","メール","tsunoda@maruu.co.jp","03-6904-8141","03-5628-3722","対象外(経理)","","経理窓口。市場担当(京葉/下館)へ送付済なら不要"],
]
# FAX先 記入用の空行
for i in range(6, 14):
    rows.append([i,"（FAX先の市場名を記入）","","市場","FAX","","","","FAX予定","","メール無しの市場はここに記入しFAX送付"])

r = hrow + 1
for data in rows:
    for c, v in enumerate(data, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = border
        cell.alignment = left if c in (2,6,11) else center
        cell.font = Font(size=9)
    r += 1

# 状況プルダウン
dv = DataValidation(type="list", formula1='"未着手,下書き作成済,送信済,FAX予定,FAX済,対象外(経理),保留"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"I{hrow+1}:I{r-1}")

widths = [4,26,12,8,8,30,15,15,14,11,34]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A5"
ws.row_dimensions[1].height = 22

wb.save(OUT)
print("checklist saved:", OUT)
